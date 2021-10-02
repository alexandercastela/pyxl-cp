
"""
Created on Tue Dec 10 15:52:06 2019

@author: jherrera
Script was last updated on: 01-08-2020 (JH)

Updates-Check excel file in folder for updates

Objective:
This script was developed with the assumption that the file where data is going to be copied from and the
file where data is going to be pasted to already exists. For now, this script is unable to create new workbooks 
or sheets.
Script will be updated and new features will be added in the future.
How it works: 
    Widget opens prompting user to enter the names of:
        1. The file location and sheet name to copy data from
        2. The file location and sheet name to paste copied data to
    Widget includes buttons to:
        1. Clear-clears all user input
        2. Run-script copies and pastes data based on user input
        3. Quit-ends the script when user is done
        4. Check button-displays sheetnames of files
Updates as of 09092021 
1. Listbox to show a list of sheet names from both excel workbooks, 
"""
import pandas as pd 
from openpyxl import load_workbook #NEED THIS MODULE IN ORDER TO COPY NEW DATA WITHOUT DELETING PREVIOUS DATA
import openpyxl as xl
from tkinter import *
from tkinter import filedialog
import tkinter as tk
from tkinter.messagebox import *
from tkinter import Listbox #09092021 - added to use listbox
import os
import sys
#For handling a "module" exception
from openpyxl.utils.exceptions import InvalidFileException
 
m = tk.Tk()
    #SETS WIDGET WINDOW SIZE WxH
m.geometry("")
m.title('Copy and Paste Widget')

"""ENTRIES FOR USER INPUT""" #09092021 - changed variable names & updated comments
#COPYING FROM FILE LOCATION-GETS INPUT
# workbookcopy=tk.Entry(m) #PREV e1
# workbookcopy.grid(row=0, column=1) #prev e1
#COPYING FROM SHEET NAME-GETS INPUT


#PASTING TO FILE LOCATION-GETS INPUT
# workbookpaste=tk.Entry(m) #prev e2
# workbookpaste.grid(row=0, column=3) #prev e2
#PASTING TO SHEET NAME-GETS INPUT


"""FUNCTION TO PRINT USER INPUT"""
def copyPaste():
#     pasteto = workbookpaste.get() #FILE TO PASTE DATA TO --> 09092021 - prev e2
#     copyfrom = workbookcopy.get() #FILE TO COPY DATA FROM --> 09092021 - prev e1
# #     copydata = pd.read_excel(copyfrom, sheet_name= e5.get()) 
#     book = load_workbook(pasteto)
#     writer = pd.ExcelWriter(pasteto, engine = 'openpyxl') 
#     writer.book = book
#     writer.sheets = {ws.title: ws for ws in book.worksheets} 
#     sheetname = sheetpaste.get() #GETS THE NAME OF THE SHEET TO PASTE TO
#     copydata.to_excel(writer, sheet_name = sheetname,
#                       startrow = writer.sheets[sheetname].max_row, index = False, header = False, ) 
# #    if var.get() == 1:
# #        print("Check button value: %s" % var.get())
#     book.creat_sheet('New Sheet', 0)
#     writer.save() #SAVES FILE
#     #writer.saveAs("New File")
    print("SCRIPT IS DONE RUNNING") #MESSAGE IS PRINTED TO THE IPYTHON CONSOLE
    #os.system(pasteto)

#list sheetnames for user selection

"""FUNCTION TO CLEAR USER INPUT WHEN "Clear" BUTTON IS PUSHED"""
def clearText(): #09092021 - updated variable names to reflect change
    print("No was chosen. Script is finished")
    # workbookcopy.delete(0, 'end')
    # workbookpaste.delete(0, 'end')

   

"""FUNCTION TO END THE PROGRAM WHEN "Quit" BUTTON IS PUSHED"""
def endScript():
    result = askyesno("Quit Prompt", "Are you sure you want to end the script?")
    if result == True:
        print("script has ended")
        m.destroy()

"""Function to show the list of sheetnames within each file"""
def getSheetNames(a):
    print("file1", a)
    file1 = xl.load_workbook(a)
    filelist1 = tk.StringVar(value=file1.sheetnames)
    listbox1 = tk.Listbox(
        m, 
        listvariable=filelist1,
        height=6,
        selectmode=SINGLE
    )
    listbox1.grid(row=3)
    d=listbox1.curselection()
    print(d)
    # testPrint(d)
    print("first file", file1.sheetnames)
    # return file1.sheetnames

def getSheetNames2(b):
    print("file 2", b)
    file2 = xl.load_workbook(b)
    # getSheetNames2.test = file2
    filelist2 = tk.StringVar(value=file2.sheetnames)
    listbox2 = tk.Listbox(
        m, 
        listvariable=filelist2,
        height=6,
        selectmode=SINGLE
    )
    listbox2.grid(row=3, column=3)
    d=listbox2.curselection()
    print(d)
# print("printing outside function", getSheetNames())

#     #try:
# #        raise OSError('Testing raising an OSError')
#         # if var1.get() == 1:
#             pasteto = workbookpaste.get() #09092021 - prev e2
#             copyfrom = workbookcopy.get() #09092021 - prev e1
#             file1 = xl.load_workbook(pasteto)
#             file2 = xl.load_workbook(copyfrom)
#         #    test1 = file1.sheetnames
#             print('\n'.join(file1.sheetnames)) 
#             print(os.path.basename(pasteto)) 
#             # print(file1.sheetnames)   
#             #LISTBOX WITH LIST OF SHEETNAMES - 09122021
#             # wbcopy = pd.ExcelFile(file1)
#             # wbpaste = pd.ExcelFile(file2)
#             copylist = tk.StringVar(value=file1.sheetnames)
#             pastelist = tk.StringVar(value = file2.sheetnames)
#             listbox2 = tk.Listbox (
#                 m, 
#                 listvariable=pastelist,
#                 height=6
#             )
#             listbox2.grid(row=2, column=1)
#             listbox = tk.Listbox(
#                 m,
#                 listvariable=copylist,
#                 height=6
#             )
#             listbox.grid(row=2, column=2)

              
            #DISPLAYS SHEETNAMES FROM FILE DATA IS BEING PASTED TO
            # tk.Label(m, text ="List of sheetnames from (pasting to) file, %s" % os.path.basename(pasteto)).grid(sticky = tk.S)
            # tk.Label(m, text = '\n'.join(file1.sheetnames)).grid(sticky = tk.S)
            # #DISPLAYS SHEETNAMES FROM FILE DATA IS BEING COPIED FROM
            # tk.Label(m, text ="List of sheetnames from (copying from) file, %s" % os.path.basename(copyfrom)).grid(sticky = tk.S)
            # tk.Label(m, text = '\n'.join(file2.sheetnames)).grid(sticky = tk.S)
        # elif var1.get() == 0:
            # tk.Label(m, text = " ").grid(sticky = tk.S)
        #     except InvalidFileException as e:
        # fileError = messagebox.showerror('File Exception Error',
        #         'Please make sure that both file names have been entered. \n\nFile names must be formatted as such: \n C:/Users/username.BC/Desktop/example_file.xlsx')
#    except OSError as e:
#        print("GOT AN OSERROR -_-: " + repr(e))


    # return file2
    # print("second file", file2.sheetnames)        
# getSheetNames2()
# print("printing from outside function", getSheetNames2.test)
def testPrint(str, c):
    print("printing from testing function", c)

#FUNC1TION TO CREATE A NEW SHEET IN AN ALREADY EXISTING EXCEL FILE (1/13/20)
#def newSheet():
#    print("gonna save as a new sheet")
    #book.creat_sheet()
#CREATE AND SAVE A NEW SHEET IN EXISTING EXCEL FILE-BASED ON COPIED DATA (1/13/20)
#tk.Label(m, text="Would you like to create a new sheet?").grid(row = 7, column = 0)
##tk.Button(m, text="Yes", command = newSheet).grid(row = 7, column =1, pady = 4)
#tk.Checkbutton(m, text = "Yes", variable = var, onvalue =1, command = copyPaste).grid(row = 7, column =1)
#print("Check button value: %s" % var.get())

def openFile():
    open_file = filedialog.askopenfilename(title="Select file", filetypes=(("Execel files", ".xlsx .xls"),))
    # print(os.path.basename(open_file))
    tk.Label(m, text=os.path.basename(open_file)).grid(row=2)
    getSheetNames(open_file)

def openFile2():
    open_file2 = filedialog.askopenfilename(title="Select file", filetypes=(("Execel files", ".xlsx .xls"),))
    # print(os.path.basename(open_file2))
    tk.Label(m, text=os.path.basename(open_file2)).grid(row=2, column=3)
    getSheetNames2(open_file2)
"""LABELS""" #09092021 - updated comments
tk.Label(m, text="Choose file to copy data from:").grid(row=0) #workbookcopy, prev e1
tk.Label(m, text="Choose file to paste data to:" ).grid(row=0, column=3) #workbookpaste, e3
tk.Button(m, text='Open File', command=openFile).grid(row=1)
tk.Button(m, text='Open File', command=openFile2).grid(row=1, column=3)
"""BUTTONS"""
tk.Button(m, text='Clear', command = clearText).grid(row=8, column=0)
tk.Button(m, text='Run', command = copyPaste).grid(row=8, column=1)
tk.Button(m, text="Test Print", command=testPrint).grid(row=8, column=3)
#tk.Button(m, text='Run', command = showSheetNames).grid(row=8, column=2)
#tk.Button(m, text='Show Input', command = printInput).grid(row=8, column=0)
tk.Button(m, text='Quit', command = endScript).grid(row=8, column=2)
# tk.Button(m, text='Show Sheets', command = getSheetNames).grid(row=8, column=3)
"""CHECKBOX - TO GET/DISPLAY SHEET NAMES FROM FILES""" #09092021 - commented out section
# var1 = IntVar() #variable for checkbutton-getSheetNames func
# tk.Label(m, text = "Display files sheetnames? ").grid(sticky = tk.S)
# tk.Checkbutton(m, text = "Yes", variable = var1, onvalue = 1, offvalue = 0, command = getSheetNames).grid(sticky = tk.S)



# pasteto = workbookpaste.get() #09092021 - prev e2
# copyfrom = workbookcopy.get() #09092021 - prev e1
# file1 = xl.load_workbook(pasteto)
# file2 = xl.load_workbook(copyfrom)
# copylist = tk.StringVar(value=file1.sheetnames)
# pastelist = tk.StringVar(value = file2.sheetnames)
# listbox2 = tk.Listbox (
#     m, 
#     listvariable=pastelist,
#     height=6
#     )
# listbox2.grid(row=2, column=1)
# listbox = tk.Listbox(
#     m,
#     listvariable=copylist,
#     height=6
#     )
# listbox.grid(row=2, column=2)

# try:
#     from exception import myexception
# except Exception as e:
#     print 
m.mainloop()
