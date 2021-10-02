
"""
Created on Tue Dec 10 15:52:06 2019

@author: jherrera
Script was last updated on: 01-08-2020 (JH)

Updates-Check excel file in folder for updates

Objective:
This script was developed with the assumption that the file where data is going to be copied from and the
file where data is going to be pasted to already exists. For now, this script is unable to create new workbooks 
or sheets.
Script will be updated and new features will be added in the future.
How it works: 
    Widget opens prompting user to enter the names of:
        1. The file location and sheet name to copy data from
        2. The file location and sheet name to paste copied data to
    Widget includes buttons to:
        1. Clear-clears all user input
        2. Run-script copies and pastes data based on user input
        3. Quit-ends the script when user is done
        4. Check button-displays sheetnames of files
"""
import pandas as pd 
from openpyxl import load_workbook #NEED THIS MODULE IN ORDER TO COPY NEW DATA WITHOUT DELETING PREVIOUS DATA
import openpyxl as xl
from tkinter import *
import tkinter as tk
from tkinter.messagebox import *
import os
import sys



#For handling a "module" exception
from openpyxl.utils.exceptions import InvalidFileException

m = tk.Tk()
    #SETS WIDGET WINDOW SIZE WxH
#m.geometry('800x650')
m.title('Copy and Paste Widget')



"""FUNCTION TO PRINT USER INPUT"""
def copyPaste():
#     pasteto = e2.get() #FILE TO PASTE DATA TO
#     copyfrom = e1.get() #FILE TO COPY DATA FROM
#     copydata = pd.read_excel(copyfrom, sheet_name= e5.get()) 
#     book = load_workbook(pasteto)
#     writer = pd.ExcelWriter(pasteto, engine = 'openpyxl') 
#     writer.book = book
#     writer.sheets = {ws.title: ws for ws in book.worksheets} 
#     sheetname = e6.get() #GETS THE NAME OF THE SHEET TO PASTE TO
#     copydata.to_excel(writer, sheet_name = sheetname,
#                       startrow = writer.sheets[sheetname].max_row, index = False, header = False, ) 
# #    if var.get() == 1:
# #        print("Check button value: %s" % var.get())
#     book.creat_sheet('New Sheet', 0)
#     writer.save() #SAVES FILE
#     #writer.saveAs("New File")
    print("SCRIPT IS DONE RUNNING") #MESSAGE IS PRINTED TO THE IPYTHON CONSOLE
    #os.system(pasteto)

#list sheetnames for user selection

"""FUNCTION TO CLEAR USER INPUT WHEN "Clear" BUTTON IS PUSHED"""
def clearText():
    print("No was chosen. Script is finished")
    e1.delete(0, 'end')
    e2.delete(0, 'end')
    e5.delete(0, 'end')
    e6.delete(0, 'end')

"""FUNCTION TO END THE PROGRAM WHEN "Quit" BUTTON IS PUSHED"""
def endScript():
    result = askyesno("Quit Prompt", "Are you sure you want to end the script?")
    if result == True:
        print("script has ended")
        m.destroy()

"""Function to show the list of sheetnames within each file"""
def getSheetNames():
    #try:
#        raise OSError('Testing raising an OSError')
        if var1.get() == 1:
            pasteto = e2.get()
            copyfrom = e1.get()
            file1 = xl.load_workbook(pasteto)
            file2 = xl.load_workbook(copyfrom)
        #    test1 = file1.sheetnames
            print('\n'.join(file1.sheetnames))
            print(os.path.basename(pasteto))
            #print(file2.sheetnames)      
            #DISPLAYS SHEETNAMES FROM FILE DATA IS BEING PASTED TO
            tk.Label(m, text ="List of sheetnames from (pasting to) file, %s" % os.path.basename(pasteto)).grid(sticky = tk.S)
            tk.Label(m, text = '\n'.join(file1.sheetnames)).grid(sticky = tk.S)
            #DISPLAYS SHEETNAMES FROM FILE DATA IS BEING COPIED FROM
            tk.Label(m, text ="List of sheetnames from (copying from) file, %s" % os.path.basename(copyfrom)).grid(sticky = tk.S)
            tk.Label(m, text = '\n'.join(file2.sheetnames)).grid(sticky = tk.S)
        elif var1.get() == 0:
            tk.Label(m, text = " ").grid(sticky = tk.S)
        #     except InvalidFileException as e:
        # fileError = messagebox.showerror('File Exception Error',
        #         'Please make sure that both file names have been entered. \n\nFile names must be formatted as such: \n C:/Users/username.BC/Desktop/example_file.xlsx')
#    except OSError as e:
#        print("GOT AN OSERROR -_-: " + repr(e))
         
       
#FUNCTION TO CREATE A NEW SHEET IN AN ALREADY EXISTING EXCEL FILE (1/13/20)
#def newSheet():
#    print("gonna save as a new sheet")
    #book.creat_sheet()
#CREATE AND SAVE A NEW SHEET IN EXISTING EXCEL FILE-BASED ON COPIED DATA (1/13/20)
#tk.Label(m, text="Would you like to create a new sheet?").grid(row = 7, column = 0)
##tk.Button(m, text="Yes", command = newSheet).grid(row = 7, column =1, pady = 4)
#tk.Checkbutton(m, text = "Yes", variable = var, onvalue =1, command = copyPaste).grid(row = 7, column =1)
#print("Check button value: %s" % var.get())

"""LABELS"""
tk.Label(m, text="Enter file location to copy data from:").grid(row=0) #e1
tk.Label(m, text="Enter sheet name to copy from:").grid(row=1) #e2
tk.Label(m, text="Enter file location to paste data to:" ).grid(row=0, column=2) #e3
tk.Label(m, text="Enter sheet name to paste to:").grid(row=1, column=2) #e4

"""BUTTONS"""
tk.Button(m, text='Clear', command = clearText).grid(row=8, column=1)
tk.Button(m, text='Run', command = copyPaste).grid(row=8, column=2)
#tk.Button(m, text='Run', command = showSheetNames).grid(row=8, column=2)
#tk.Button(m, text='Show Input', command = printInput).grid(row=8, column=0)
tk.Button(m, text='Quit', command = endScript).grid(row=8, column=3)

"""CHECKBOX - TO GET/DISPLAY SHEET NAMES FROM FILES"""
var1 = IntVar() #variable for checkbutton-getSheetNames func
tk.Label(m, text = "Display files sheetnames? ").grid(sticky = tk.S)
tk.Checkbutton(m, text = "Yes", variable = var1, onvalue = 1, offvalue = 0, command = getSheetNames).grid(sticky = tk.S)

"""ENTRIES FOR USER INPUT"""
#COPYING FROM FILE LOCATION-GETS INPUT
e1=tk.Entry(m) #PREV e1
e1.grid(row=0, column=1)
#COPYING FROM SHEET NAME-GETS INPUT
e5=tk.Entry(m)
e5.grid(row=1, column=1)

#PASTING TO FILE LOCATION-GETS INPUT
e2=tk.Entry(m)
e2.grid(row=0, column=3)
#PASTING TO SHEET NAME-GETS INPUT
e6=tk.Entry(m)
e6.grid(row=1, column=3)

# try:
#     from exception import myexception
# except Exception as e:
#     print 
m.mainloop()
